from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class StockItem:
    code: str
    name: str


def normalize_stock_code(value: object) -> str:
    """Return a six digit A-share stock code."""
    if value is None:
        raise ValueError("stock code is empty")
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    text = text.zfill(6)
    if not text.isdigit() or len(text) != 6:
        raise ValueError(f"invalid stock code: {value!r}")
    return text


def to_ts_code(code: str) -> str:
    """Convert a six digit A-share code to a Tushare-style exchange code."""
    normalized = normalize_stock_code(code)
    if normalized.startswith(("60", "68", "90")):
        return f"{normalized}.SH"
    if normalized.startswith(("00", "30", "20")):
        return f"{normalized}.SZ"
    if normalized.startswith(("43", "83", "87", "92")):
        return f"{normalized}.BJ"
    raise ValueError(f"cannot infer exchange for stock code: {code!r}")


def to_akshare_symbol(code: str) -> str:
    """AKShare stock_zh_a_hist expects the plain six digit code."""
    return normalize_stock_code(code)


def read_stock_pool(path: str | Path, sheet_name: str | None = None) -> list[StockItem]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        try:
            code_index = headers.index("代码")
            name_index = headers.index("名称")
        except ValueError as exc:
            raise ValueError("stock pool must contain columns: 代码, 名称") from exc

        items: list[StockItem] = []
        seen: set[str] = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or row[code_index] in (None, ""):
                continue
            code = normalize_stock_code(row[code_index])
            if code in seen:
                continue
            name = str(row[name_index]).strip() if row[name_index] is not None else ""
            items.append(StockItem(code=code, name=name))
            seen.add(code)
        return items
    finally:
        workbook.close()


def select_stock(items: Iterable[StockItem], code: str) -> StockItem:
    normalized = normalize_stock_code(code)
    for item in items:
        if item.code == normalized:
            return item
    return StockItem(code=normalized, name="")
